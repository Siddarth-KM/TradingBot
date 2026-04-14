#!/usr/bin/env python3
"""Agent 2: Excel Writer — MODE A"""

import shutil, sys, logging
from datetime import datetime, timedelta
from pathlib import Path
import openpyxl
from openpyxl.styles import Font

EXCEL_PATH = r"C:\Users\sidda\OneDrive\Documents\InvestmentTracker.xlsx"
LOG_PATH   = Path("logs/excel_week37.log")

# Logging — stdout always; file if writable
log = logging.getLogger("excel_writer")
log.setLevel(logging.INFO)
log.addHandler(logging.StreamHandler(sys.stdout))
try:
    log.addHandler(logging.FileHandler(str(LOG_PATH), mode='w'))
except OSError:
    log.warning("Could not open log file — logging to stdout only.")

# ── Signals (VM-fetched, field-mapped) ───────────────────────────────────────
signals = [
    {"ticker":"LGIH","pred":0.12794,"probability":63.5,"buy_price":39.52, "sell_limit":42.73},
    {"ticker":"MTSI","pred":0.11337,"probability":66.0,"buy_price":263.63,"sell_limit":283.36},
    {"ticker":"CCS", "pred":0.10743,"probability":66.0,"buy_price":60.32, "sell_limit":64.6},
    {"ticker":"WLK", "pred":0.1072, "probability":66.5,"buy_price":120.15,"sell_limit":128.71},
    {"ticker":"INTC","pred":0.09347,"probability":58.5,"buy_price":65.18, "sell_limit":68.74},
    {"ticker":"LEN", "pred":0.08124,"probability":59.0,"buy_price":89.79, "sell_limit":94.09},
    {"ticker":"HII", "pred":0.08024,"probability":66.5,"buy_price":394.46,"sell_limit":415.51},
    {"ticker":"FANG","pred":0.06613,"probability":59.5,"buy_price":189.1, "sell_limit":196.54},
]
GENERATED_AT = "2026-04-14T07:41:47.014348"
log.info(f"Loaded {len(signals)} signals (generated_at={GENERATED_AT})")

# ── Step 2: week_number and date_range ───────────────────────────────────────
week_number = 37
signal_ts        = datetime.fromisoformat(GENERATED_AT)
monday           = signal_ts - timedelta(days=signal_ts.weekday())
tuesday          = monday + timedelta(days=1)
following_monday = monday + timedelta(days=7)
date_range = f"{tuesday.month}/{tuesday.day} - {following_monday.month}/{following_monday.day}"
log.info(f"Week {week_number} | date_range: {date_range}")

# ── Step 3: Starting capital ──────────────────────────────────────────────────
wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb.active

starting_capital = None
for row in ws.iter_rows():
    a_val = row[0].value
    k_val = row[10].value
    if isinstance(a_val, str) and a_val.strip() == "Total":
        if k_val not in (None, ""):
            try:
                starting_capital = float(k_val)
            except (TypeError, ValueError):
                pass

if starting_capital is None:
    log.error("No prior Total row with non-blank Final Amount — cannot write week block.")
    sys.exit(1)

log.info(f"Starting capital: ${starting_capital:,.2f}")

# ── Step 4: Backup ────────────────────────────────────────────────────────────
backup_path = EXCEL_PATH.replace('.xlsx', f'_backup_{datetime.today().strftime("%Y%m%d")}.xlsx')
shutil.copy2(EXCEL_PATH, backup_path)
log.info(f"Backup: {backup_path}")

# ── Step 5: Append week block ─────────────────────────────────────────────────
# Find last row with content
last_content_row = 0
for r in range(ws.max_row, 0, -1):
    if any(ws.cell(row=r, column=c).value not in (None, "") for c in range(1, 12)):
        last_content_row = r
        break

first_empty_row = last_content_row + 2   # blank gap
data_start      = first_empty_row + 2    # skip header + col-header rows

# Week header
ws.cell(row=first_empty_row, column=1).value = f"Week {week_number}"
ws.cell(row=first_empty_row, column=2).value = date_range
ws.cell(row=first_empty_row, column=1).font  = Font(bold=True)

# Column headers
col_hdrs = ["Ticker","Timeframe","Pred","Probability","Buy Price",
            "Sell Limit","Final Sell","Percent Return","Shares Purchased",
            "Start Amount","Final Amount"]
for ci, hdr in enumerate(col_hdrs, 1):
    ws.cell(row=first_empty_row+1, column=ci).value = hdr
    ws.cell(row=first_empty_row+1, column=ci).font  = Font(bold=True)

# Data rows
for i, sig in enumerate(signals):
    r = data_start + i
    ws.cell(row=r, column=1).value  = sig["ticker"]
    ws.cell(row=r, column=2).value  = 5
    ws.cell(row=r, column=3).value  = sig["pred"]
    ws.cell(row=r, column=4).value  = sig["probability"]
    ws.cell(row=r, column=5).value  = sig["buy_price"]
    ws.cell(row=r, column=6).value  = sig["sell_limit"]
    ws.cell(row=r, column=8).value  = f'=IF(G{r}="","",(G{r}-E{r})/E{r})'
    ws.cell(row=r, column=11).value = f'=IF(OR(I{r}="",G{r}=""),"",I{r}*G{r})'

total_r = data_start + len(signals)
ws.cell(row=total_r,   column=1).value = "Total";             ws.cell(row=total_r,   column=1).font = Font(bold=True)
ws.cell(row=total_r,   column=2).value = 5
ws.cell(row=total_r+1, column=1).value = "S&P 500(Adjusted)"; ws.cell(row=total_r+1, column=2).value = 5
ws.cell(row=total_r+2, column=1).value = "S&P 500";           ws.cell(row=total_r+2, column=2).value = 5

metrics = [
    ("Cum. Return",         "YTD Cum. Return"),
    ("Sharpe Ratio (Cum.)", "YTD Sharpe Ratio"),
    ("Max Drawdown (Cum.)", "YTD Max Drawdown"),
    ("Calmar Ratio (Cum.)", "YTD Calmar Ratio"),
    ("Sortino Ratio (Cum.)","YTD Sortino Ratio"),
    ("Info. Ratio (Cum.)",  "YTD Info. Ratio"),
    ("Win Rate",            "YTD Win Rate"),
    ("Alpha vs SPY (Cum.)", "YTD Alpha vs SPY"),
    ("Payoff Ratio (Cum.)", "Payoff Ratio (YTD)"),
    ("Profit Factor (Cum.)","Profit Factor (YTD)"),
]
for mi, (la, lc) in enumerate(metrics):
    mr = total_r + 3 + mi
    ws.cell(row=mr, column=1).value = la
    ws.cell(row=mr, column=3).value = lc

wb.save(EXCEL_PATH)
log.info("Workbook saved.")

# ── Step 6: Confirm ───────────────────────────────────────────────────────────
print(f"""
Week {week_number} open positions written to Excel:
  {len(signals)} tickers | Date range: {date_range}
  Buy prices from last_close (placeholders — correct after trading)
  Formulas written: H (Percent Return) and K (Final Amount) auto-calculate on sell price entry
  File: {EXCEL_PATH}
""")
